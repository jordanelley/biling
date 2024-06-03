# from openpyxl import Workbook
import openpyxl

path = "./Data/Monthly Checking _ Queenstown Lakes District Council 20240506.xlsx"

wb_obj = openpyxl.load_workbook(path)

wb_obj.create_sheet("new")

checking_report = wb_obj['TotalComplete']
new_sheet = wb_obj['new']

cell_obj = checking_report.cell(1, 1)

for cell in checking_report['U:U']:
    new_sheet.cell(cell.row, 1, cell.value)

for cell in checking_report['B:B']:
    new_sheet.cell(cell.row, 2, cell.value)

new_sheet.cell(1, 3, "priority")
new_sheet.cell(1, 4, "target")
new_sheet.cell(1, 5, "Activity")
new_sheet.cell(1, 6, "Team")
new_sheet.cell(1, 7, "QLDC contract")

for cell in checking_report['I:I']:
    # print('Printing from ' + str(cell.column) + str(cell.row))
    new_sheet.cell(cell.row, 8, cell.value)

for cell in checking_report['F:F']:
    # print('Printing from ' + str(cell.column) + str(cell.row))
    new_sheet.cell(cell.row, 9, cell.value)

for cell in checking_report['D:D']:
    # asset feature DB
    # print('Printing from ' + str(cell.column) + str(cell.row))
    new_sheet.cell(cell.row, 10, cell.value)

for cell in checking_report['D:D']:
    # asset feature DB
    new_sheet.cell(cell.row, 10, cell.value)

for cell in checking_report['D:D']:
    # asset feature DB
    new_sheet.cell(cell.row, 11, cell.value)

for cell in checking_report['D:D']:
    # asset feature DB
    new_sheet.cell(cell.row, 12, cell.value)


for cell in checking_report['O:O']:
    # asset feature DB
    new_sheet.cell(cell.row, 13, cell.value)

new_sheet.cell(1, 14, "Ward")

id_nums = checking_report['C:H']

for i in range(len(id_nums[0])):
    val = str(id_nums[0][i].value) + " " + str(id_nums[4][i].value or '')
    new_sheet.cell(i+1, 15, val)

for cell in checking_report['J:J']:
    # asset feature DB
    new_sheet.cell(cell.row, 16, cell.value)

new_sheet.cell(1, 17, "Cost breakdown")
new_sheet.cell(1, 18, "QUote no")
new_sheet.cell(1, 19, "scheduled dates")
new_sheet.cell(1, 20, "Notes/Updates")

for cell in checking_report['N:N']:
    new_sheet.cell(cell.row, 21, cell.value)
new_sheet.cell(1, 22, "Post Works")







wb_obj.save('test.xlsx')


